from openpyxl import Workbook
from datetime import datetime
from openpyxl import load_workbook
import csv

wb = load_workbook('lista_de_precios.xlsx')
sheets = wb.sheetnames
valores_precio = []
for i in range(len(sheets)):
    sheet = wb[sheets[i]]
    filas = sheet.rows
    for fila in filas:
        descripcion = fila[3].value
        if descripcion == 'Ruta Caribe':
            valores_precio.append(fila[5].value)
print(valores_precio)
        


