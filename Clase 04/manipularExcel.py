from openpyxl import Workbook
from datetime import datetime
wb = Workbook()

sheet = wb.active

#Escribir una celda determinada
sheet['A1'] = 42

sheet.append([1,'valor',341])

sheet['A3'] = datetime.now()

#Metodo alternativo para escribir la celda
sheet.cell(6,4,'Alternativo')


wb.save('ejemplo.xlsx')

###Lectura Libros excel###
from openpyxl import load_workbook

wb = load_workbook('tabla.xlsx')
#sheet = wb['Hoja1']

###Metdo alternativo
sheets = wb.sheetnames
sheet = wb[sheets[0]]

#Capturar la columna entera
columna = sheet['A']
for valor in columna:
    print(valor.value)
    
#Capturar Fila
fila = sheet['2']
for valor in fila:
    print(valor.value)
    
#Recorrer filas
filas = sheet.rows
for fila in filas:
    print('Filas: ')
    for celda in fila:
        print(celda.value)
        
#Recorrer columna
columnas = sheet.columns
for column in columnas:
    print('Columnas: ')
    for celda in column:
        print(celda.value)
        
import csv

with open('data.csv','w',encoding='utf-8',newline='') as salida:
    salida2 = csv.writer(salida)
    filas = sheet.rows
    for fila in filas:
        lista = [fila[x].value for x in range(len(fila))]
        salida2.writerow(lista)
        
        
        
        
        
        
        
        
        
        